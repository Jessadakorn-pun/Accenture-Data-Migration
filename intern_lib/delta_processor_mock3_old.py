# delta_processor_mock3_final.py

import os
import time
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side


class Timer:
    """Context manager for timing a code block."""

    def __init__(self, name: str):
        """
        Initialize the timer with a descriptive name.
        :param name: A label to identify this timed block in output.
        """
        self.name = name

    def __enter__(self):
        """Start timing and print a start message."""
        self.start = time.time()
        print(f"[START] {self.name}")
        return self

    def __exit__(self, exc_type, exc, tb):
        """
        Stop timing, compute elapsed time, and print a done message.
        """
        elapsed = time.time() - self.start
        m, s = divmod(elapsed, 60)
        print(f"[DONE]  {self.name} — {int(m)}m {s:.2f}s\n")


class DeltaProcessor:
    """
    Orchestrates delta detection (Delete, New, Change) across multiple Excel workbooks.
    """

    def __init__(self, master_path: str):
        """
        Load master workbook and read configuration parameters.
        """
        if not os.path.isfile(master_path):
            raise FileNotFoundError(f"Master workbook not found: {master_path}")
        self.master_path = master_path
        self.params = {}
        self.file_list = None
        self.exceptions = []
        self._pink = PatternFill(start_color="F2CEF0",
                                 end_color="F2CEF0",
                                 fill_type="solid")
        self._load_master()

    def _load_master(self):
        """
        Read the 'Parameters', 'File List', and 'Exception Sheet Name' sheets
        from the master workbook into memory.
        """
        p = pd.read_excel(self.master_path,
                          sheet_name="Parameters",
                          header=None, index_col=0)[1]
        p.index = p.index.astype(str).str.strip()
        self.params = {
            'first_row':         int(p['firstRowOfData']),
            'mock_col':          int(p['mockColumn']),
            'delta_col':         int(p['deltaIndicatorColumn']),
            'key_col':           int(p['concatenatedKeyColumn']),
            'status_prev_col':   int(p['statusFromPreviousMockColumn']),
            'record_limit':      int(p['recordLimit']),
            'header_row':        int(p['headerRow']),
            'start_compare_col': int(p['startColumnCheckData'])
        }
        self.file_list = pd.read_excel(self.master_path, sheet_name="File List")
        exc = pd.read_excel(self.master_path, sheet_name="Exception Sheet Name").iloc[:, 0]
        self.exceptions = exc.dropna().astype(str).str.strip().tolist()

    def run(self):
        """
        Execute delta processing for each workbook listed in the 'File List'.
        """
        with Timer("TOTAL RUN"):
            for _, row in self.file_list.iterrows():
                path, fname = row.iloc[0], row.iloc[1]
                full_path = os.path.join(path, fname)
                if os.path.exists(full_path):
                    self._process_file(full_path)

    def _process_file(self, file_path: str):
        """
        Process a single Excel file: iterate through sheets, apply deltas, and save.
        """
        wb_name = os.path.basename(file_path)
        with Timer(f"Workbook: {wb_name}"):
            wb = load_workbook(file_path)
            p = self.params
            fr = p['first_row']

            for ws_name in wb.sheetnames:
                if ws_name in self.exceptions:
                    continue

                with Timer(f"Sheet: {ws_name}"):
                    ws = wb[ws_name]

                    # 1) Backup
                    with Timer("backup"):
                        backup = wb.copy_worksheet(ws)
                        backup.title = f"{ws_name}_backup"

                    # 2) Load into DataFrame
                    with Timer("load df"):
                        df = pd.read_excel(
                            file_path,
                            sheet_name=ws_name,
                            header=fr - 2,
                            dtype=str
                        )

                    # 3) Validate mock flags
                    with Timer("validate mock flags"):
                        mc = p['mock_col'] - 1
                        if not df.iloc[:, mc].isin(['2', '3']).all():
                            raise ValueError(f"Invalid mock flags in {ws_name}")

                    # 4) Clear old delta indicators
                    with Timer("clear old deltas"):
                        dc = df.columns[p['delta_col'] - 1]
                        df[dc] = ''

                    # 5) Sort by key & mock
                    with Timer("sort by key/mock"):
                        kc = p['key_col'] - 1
                        df.sort_values(by=[df.columns[kc], df.columns[mc]],
                                       inplace=True)
                        df.reset_index(drop=True, inplace=True)

                    # 6) DeltaDelete
                    with Timer("delta_delete"):
                        df = self._delta_delete(df)

                    # 7) DeltaNew
                    with Timer("delta_new"):
                        df = self._delta_new(df)

                    # 8) DeltaChange
                    with Timer("delta_change"):
                        drop_idx, to_color = self._delta_change(df, ws)

                    # 9) Drop identical rows & remap highlights
                    with Timer("drop identical rows"):
                        if drop_idx:
                            df.drop(index=drop_idx, inplace=True)
                            df.reset_index(drop=True, inplace=True)
                            drops = sorted(drop_idx)
                            to_color = [
                                (r - sum(1 for d in drops if d < r), c)
                                for r, c in to_color
                            ]

                    # 10) Rewrite sheet contents
                    with Timer("rewrite sheet"):
                        if ws.max_row >= fr:
                            ws.delete_rows(fr, ws.max_row - fr + 1)
                        for row in dataframe_to_rows(df, index=False, header=False):
                            ws.append(row)

                    # 11) Restore formatting + table & font & borders
                    with Timer("restore formatting"):
                        self._restore_sheet_format(ws, wb[f"{ws_name}_backup"], fr)

                    # 12) Highlight changes
                    with Timer("highlight changes"):
                        for r, c in to_color:
                            ws.cell(row=fr + r, column=c + 1).fill = self._pink

            wb.save(file_path)
            wb.close()

    def _delta_delete(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Mark rows as 'Delete' where baseline (mock=2) keys do not appear in snapshot (mock=3).
        """
        p = self.params
        mc, kc, sc, dc = (
            p['mock_col'] - 1,
            p['key_col'] - 1,
            p['status_prev_col'] - 1,
            p['delta_col'] - 1
        )
        rng = df if p['record_limit'] == 0 else df.iloc[:p['record_limit']]
        new_keys = set(rng.loc[rng.iloc[:, mc] == '3', df.columns[kc]])
        mask = (
            (df.iloc[:, mc] == '2') &
            ~df.iloc[:, kc].isin(new_keys) &
            (df.iloc[:, sc] != 'Add')
        )
        df.loc[mask, df.columns[dc]] = 'Delete'
        return df

    def _delta_new(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Mark rows as 'New' where snapshot (mock=3) keys do not appear in baseline (mock=2).
        """
        p = self.params
        mc, kc, dc = (
            p['mock_col'] - 1,
            p['key_col'] - 1,
            p['delta_col'] - 1
        )
        rng = df if p['record_limit'] == 0 else df.iloc[:p['record_limit']]
        old_keys = set(rng.loc[rng.iloc[:, mc] == '2', df.columns[kc]])
        mask = (df.iloc[:, mc] == '3') & ~df.iloc[:, kc].isin(old_keys)
        df.loc[mask, df.columns[dc]] = 'New'
        return df

    def _delta_change(self, df: pd.DataFrame, ws) -> tuple[list[int], list[tuple[int,int]]]:
        """
        Identify and mark changes where the same key appears with different data.
        """
        p = self.params
        mc, kc, dc = (
            p['mock_col'] - 1,
            p['key_col'] - 1,
            p['delta_col'] - 1
        )
        is_new = df.iloc[:, mc] == '3'
        is_old = df.iloc[:, mc].shift(1) == '2'
        same = df.iloc[:, kc] == df.iloc[:, kc].shift(1)
        pair = is_new & is_old & same

        hdr = p['header_row']
        valid = [
            j for j in range(p['start_compare_col'] - 1, df.shape[1])
            if not any(
                x in str(ws.cell(row=hdr, column=j + 1).value or '').lower()
                for x in ['__', 'tobe', 'to be', 'to-be', 'to_be']
            )
        ]

        drop_idx, to_color = [], []
        for i in range(len(df) - 1, 0, -1):
            if not pair.iat[i]:
                continue
            new_vals = df.iloc[i, valid].fillna('').astype(str)
            old_vals = df.iloc[i - 1, valid].fillna('').astype(str)
            diffs = new_vals != old_vals
            if not diffs.any():
                drop_idx.append(i)
            else:
                df.iat[i, dc] = df.iat[i - 1, dc] = 'Change'
                for idx, changed in enumerate(diffs):
                    if changed:
                        to_color += [(i, valid[idx]), (i - 1, valid[idx])]
        return drop_idx, to_color

    def _restore_sheet_format(self, live_ws, backup_ws, first_row: int):
        """
        1) Copy column widths & row heights
        2) Copy header styles
        3) Apply Leelawadee font size 10 and thin border to loaded data
        """
        # 1) Column widths
        for col_letter, dim in backup_ws.column_dimensions.items():
            if dim.width:
                live_ws.column_dimensions[col_letter].width = dim.width

        # 2) Row heights
        for row_idx, dim in backup_ws.row_dimensions.items():
            if dim.height:
                live_ws.row_dimensions[row_idx].height = dim.height

        # 3) Header‐row styles
        hdr = first_row - 1
        for cell in backup_ws[hdr]:
            tgt = live_ws.cell(row=hdr, column=cell.column)
            for attr in (
                'font', 'border', 'fill',
                'number_format', 'protection', 'alignment'
            ):
                try:
                    setattr(tgt, attr, getattr(cell, attr))
                except:
                    pass

        # 4) Leelawadee font & thin border on data cells
        data_font = Font(name="Leelawadee", size=10)
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        max_row = live_ws.max_row
        max_col = live_ws.max_column
        for row in live_ws.iter_rows(
            min_row=first_row, max_row=max_row,
            min_col=1, max_col=max_col
        ):
            for cell in row:
                cell.font = data_font
                cell.border = border
                
if __name__ == '__main__':
    dp = DeltaProcessor(master_path="Master.xlsx")
    dp.run()
