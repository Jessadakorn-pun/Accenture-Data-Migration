import os
import time
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side

class DeltaProcessor:
    """
    DeltaProcessor to apply delta logic (Delete, New, Change) matching the VBA logic,
    comparing mock values 2 (old) and 3 (new), without the Sort Record feature.

    Includes optional mock-value validation via EnableMockNumberCheck.

    Steps:
      1) DeltaDelete  (identify old records not in new)
      2) DeltaNew     (identify new records not in old)
      3) DeltaChange  (compare old vs new pairs and highlight changes)

    Backups each sheet, then rebuilds it and applies pink highlights.
    """

    def __init__(self, master_path: str):
        if not os.path.isfile(master_path):
            raise FileNotFoundError(f"Master workbook not found: {master_path}")
        self.master_path = master_path
        self.params = {}
        self.file_list = None
        self.exceptions = []
        self._pink = PatternFill(start_color="F2CEF0",
                                 end_color="F2CEF0",
                                 fill_type="solid")
        self._load_master()

    def _load_master(self):
        p = pd.read_excel(self.master_path,
                          sheet_name="Parameters",
                          header=None, index_col=0)[1]
        p.index = p.index.astype(str).str.strip()
        self.params = {
            'first_row':             int(p['firstRowOfData']),
            'mock_col':              int(p['mockColumn']),
            'delta_col':             int(p['deltaIndicatorColumn']),
            'key_col':               int(p['concatenatedKeyColumn']),
            'status_prev_col':       int(p['statusFromPreviousMockColumn']),
            'record_limit':          int(p['recordLimit']),
            'header_row':            int(p['headerRow']),
            'start_compare_col':     int(p['startColumnCheckData']),
            'enable_mock_check':     int(p.get('EnableMockNumberCheck', 0))
        }
        self.file_list = pd.read_excel(self.master_path, sheet_name="File List")
        exc = pd.read_excel(self.master_path,
                            sheet_name="Exception Sheet Name").iloc[:,0]
        self.exceptions = exc.dropna().astype(str).str.strip().tolist()

    def run(self):
        global_start = time.time()
        for _, row in self.file_list.iterrows():
            path, fname = row.iloc[0], row.iloc[1]
            full_path = os.path.join(path, fname)
            if isinstance(path, str) and isinstance(fname, str) and os.path.exists(full_path):
                self._process_file(full_path)
        total = time.time() - global_start
        m, s = divmod(int(total), 60)
        print(f"\nALL WORKBOOKS DONE in {m}m {s}s")

    def _process_file(self, file_path: str):
        print(f"\nProcessing {os.path.basename(file_path)}")
        wb = load_workbook(file_path)
        fr = self.params['first_row']
        t0 = time.time()

        for ws_name in wb.sheetnames:
            if ws_name in self.exceptions:
                continue
            
            print(f" Sheet “{ws_name}”:")
            print(f" Start Time : {datetime.datetime.now()}")
            
            
            ws = wb[ws_name]

            # Backup sheet
            t_backup_start = time.time()
            backup = wb.copy_worksheet(ws)
            backup.title = f"{ws_name}_backup"
            t_backup_end = time.time()
            print(f"  • Backup sheet : {t_backup_end - t_backup_start:.2f}s")

            # Load into DataFrame
            t_load_start = time.time()
            df = pd.read_excel(
                file_path,
                sheet_name=ws_name,
                header=fr - 2,
                dtype=str
            )
            t_load_end = time.time()
            print(f"  • Load DataFrame : {t_load_end - t_load_start:.2f}s")

            # Validate mock values if enabled
            if self.params['enable_mock_check'] == 1:
                t_mock_start = time.time()
                mc = self.params['mock_col'] - 1
                invalid = ~df.iloc[:, mc].isin(['3', '4'])
                if invalid.any():
                    rows = df.index[invalid].tolist()
                    raise ValueError(f"Incorrect mock values in rows {rows} of sheet '{ws_name}'")
                t_mock_end = time.time()
                print(f"  • Validate mock values : {t_mock_end - t_mock_start:.2f}s")
            
            # sort by [key, mock]
            t25 = time.time()
            kc = self.params['key_col'] - 1
            mc = self.params['mock_col'] - 1
            df.sort_values(by=[df.columns[kc], df.columns[mc]], inplace=True)
            df.reset_index(drop=True, inplace=True)
            print(f"  • sort: {time.time()-t25:.2f}s")

            # DeltaDelete: old (3) not in new (4)
            t_delete_start = time.time()
            df = self._delta_delete(df)
            t_delete_end = time.time()
            print(f"  •  DeltaDelete : {t_delete_end - t_delete_start:.2f}s")

            # DeltaNew: new (4) not in old (3)
            t_new_start = time.time()
            df = self._delta_new(df)
            t_new_end = time.time()
            print(f"  • DeltaNew : {t_new_end - t_new_start:.2f}s")

            # DeltaChange: compare adjacent old-new pairs
            t_change_start = time.time()
            drop_idx, to_color = self._delta_change(df, ws)
            t_change_end = time.time()
            print(f"  • DeltaChange : {t_change_end - t_change_start:.2f}s")

            # Drop identical new rows
            t_drop_start = time.time()
            if drop_idx:
                df.drop(index=drop_idx, inplace=True)
                df.reset_index(drop=True, inplace=True)
            t_drop_end = time.time()
            print(f"  • Drop identical rows : {t_drop_end - t_drop_start:.2f}s")

            # Remap highlight coordinates after drops
            t_remap_start = time.time()
            all_drops = sorted(drop_idx)
            remapped = []
            for (r, c) in to_color:
                offset = sum(1 for dr in all_drops if dr < r)
                remapped.append((r - offset, c))
            to_color = remapped
            t_remap_end = time.time()
            print(f"  • Remap highlights : {t_remap_end - t_remap_start:.2f}s")

            # Clear existing data rows
            t_clear_start = time.time()
            num_old = ws.max_row - fr + 1
            if num_old > 0:
                ws.delete_rows(fr, amount=num_old)
            t_clear_end = time.time()
            print(f"  • Clear old rows : {t_clear_end - t_clear_start:.2f}s")

            # Write updated rows
            t_write_start = time.time()
            for row in dataframe_to_rows(df, index=False, header=False):
                ws.append(row)
            t_write_end = time.time()
            print(f"  • Write rows : {t_write_end - t_write_start:.2f}s")

            # Restore formatting
            t_restore_start = time.time()
            backup_ws = wb[f"{ws_name}_backup"]
            self._restore_sheet_format(ws, backup_ws, fr)
            t_restore_end = time.time()
            print(f"  • Restore formatting : {t_restore_end - t_restore_start:.2f}s")

            # Highlight changes in pink
            t_highlight_start = time.time()
            for r, c in to_color:
                ws.cell(row=fr + r, column=c + 1).fill = self._pink
            t_highlight_end = time.time()
            print(f"  • Highlight cells : {t_highlight_end - t_highlight_start:.2f}s \n")

        wb.save(file_path)
        print(f" → file done in {time.time()-t0:.2f}s")

    # (Other methods unchanged)

    def _delta_delete(self, df: pd.DataFrame) -> pd.DataFrame:
        p = self.params
        mc = p['mock_col'] - 1
        kc = p['key_col'] - 1
        sc = p['status_prev_col'] - 1
        dc = p['delta_col'] - 1
        rng = df if p['record_limit'] == 0 else df.iloc[:p['record_limit']]
        new_keys = set(rng.loc[rng.iloc[:, mc] == '4', df.columns[kc]])
        mask = (
            (df.iloc[:, mc] == '3') &
            (~df.iloc[:, kc].isin(new_keys)) &
            (df.iloc[:, sc] != 'Add')
        )
        df.loc[mask, df.columns[dc]] = 'Delete'
        return df

    def _delta_new(self, df: pd.DataFrame) -> pd.DataFrame:
        p = self.params
        mc = p['mock_col'] - 1
        kc = p['key_col'] - 1
        dc = p['delta_col'] - 1
        rng = df if p['record_limit'] == 0 else df.iloc[:p['record_limit']]
        old_keys = set(rng.loc[rng.iloc[:, mc] == '3', df.columns[kc]])
        mask = (df.iloc[:, mc] == '4') & (~df.iloc[:, kc].isin(old_keys))
        df.loc[mask, df.columns[dc]] = 'New'
        return df

    def _delta_change(self, df: pd.DataFrame, ws) -> tuple[list[int], list[tuple[int,int]]]:
        p = self.params
        mc = p['mock_col'] - 1
        kc = p['key_col'] - 1
        dc = p['delta_col'] - 1
        hdr = p['header_row']
        n = len(df)

        is_new = df.iloc[:, mc] == '4'
        is_old = df.iloc[:, mc].shift(1) == '3'
        same_key = df.iloc[:, kc] == df.iloc[:, kc].shift(1)
        pair = is_new & is_old & same_key

        start = p['start_compare_col'] - 1
        valid = []
        for j in range(start, df.shape[1]):
            h = str(ws.cell(row=hdr, column=j + 1).value or '').lower()
            if any(x in h for x in ('__', 'tobe', 'to be', 'to-be', 'to_be')):
                continue
            valid.append(j)

        drop_idx, to_color = [], []
        for i in range(n - 1, 0, -1):
            if not pair.iat[i]:
                continue
            new_vals = df.iloc[i, valid].fillna('').astype(str)
            old_vals = df.iloc[i - 1, valid].fillna('').astype(str)
            diffs = new_vals != old_vals
            if not diffs.any():
                drop_idx.append(i)
            else:
                df.iat[i, dc] = 'Change'
                df.iat[i - 1, dc] = 'Change'
                for idx, changed in enumerate(diffs):
                    if changed:
                        col = valid[idx]
                        to_color.extend([(i, col), (i - 1, col)])
        return drop_idx, to_color

    def _restore_sheet_format(self, live_ws, backup_ws, first_row: int):
        for col_letter, dim in backup_ws.column_dimensions.items():
            if dim.width:
                live_ws.column_dimensions[col_letter].width = dim.width
        for row_idx, dim in backup_ws.row_dimensions.items():
            if dim.height:
                live_ws.row_dimensions[row_idx].height = dim.height

        hdr = first_row - 1
        for cell in backup_ws[hdr]:
            tgt = live_ws.cell(row=hdr, column=cell.column)
            for attr in ('font', 'border', 'fill', 'number_format', 'protection', 'alignment'):
                try:
                    setattr(tgt, attr, getattr(cell, attr))
                except:
                    pass

        data_font = Font(name="Leelawadee", size=10)
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in live_ws.iter_rows(min_row=first_row, max_row=live_ws.max_row,
                                     min_col=1, max_col=live_ws.max_column):
            for cell in row:
                cell.font = data_font
                cell.border = border

if __name__ == '__main__':
    dp = DeltaProcessor(master_path="Master.xlsx")
    dp.run()
