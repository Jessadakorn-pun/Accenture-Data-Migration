import os
import time
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side

class DeltaProcessor:
    """
    DeltaProcessor to apply delta logic (Delete, New, Change) matching the VBA logic,
    comparing mock values 2 (old) and 3 (new), without the Sort Record feature.

    Includes optional mock-value validation via EnableMockNumberCheck.

    Steps:
      1) DeltaDelete  (identify old records not in new)
      2) DeltaNew     (identify new records not in old)
      3) DeltaChange  (compare old vs new pairs and highlight changes)

    Backups each sheet, then rebuilds it and applies pink highlights.
    """

    def __init__(self, master_path: str):
        if not os.path.isfile(master_path):
            raise FileNotFoundError(f"Master workbook not found: {master_path}")
        self.master_path = master_path
        self.params = {}
        self.file_list = None
        self.exceptions = []
        self._pink = PatternFill(start_color="F2CEF0",
                                 end_color="F2CEF0",
                                 fill_type="solid")
        self._load_master()

    def _load_master(self):
        p = pd.read_excel(self.master_path,
                          sheet_name="Parameters",
                          header=None, index_col=0)[1]
        p.index = p.index.astype(str).str.strip()
        self.params = {
            'first_row':             int(p['firstRowOfData']),
            'mock_col':              int(p['mockColumn']),
            'delta_col':             int(p['deltaIndicatorColumn']),
            'key_col':               int(p['concatenatedKeyColumn']),
            'status_prev_col':       int(p['statusFromPreviousMockColumn']),
            'record_limit':          int(p['recordLimit']),
            'header_row':            int(p['headerRow']),
            'start_compare_col':     int(p['startColumnCheckData']),
            'enable_mock_check':     int(p.get('EnableMockNumberCheck', 0))
        }
        self.file_list = pd.read_excel(self.master_path, sheet_name="File List")
        exc = pd.read_excel(self.master_path,
                            sheet_name="Exception Sheet Name").iloc[:,0]
        self.exceptions = exc.dropna().astype(str).str.strip().tolist()

    def run(self):
        global_start = time.time()
        for _, row in self.file_list.iterrows():
            path, fname = row.iloc[0], row.iloc[1]
            full_path = os.path.join(path, fname)
            if isinstance(path, str) and isinstance(fname, str) and os.path.exists(full_path):
                self._process_file(full_path)
        total = time.time() - global_start
        m, s = divmod(int(total), 60)
        print(f"\nALL WORKBOOKS DONE in {m}m {s}s")

    def _process_file(self, file_path: str):
        print(f"\nProcessing {os.path.basename(file_path)}")
        wb = load_workbook(file_path)
        fr = self.params['first_row']
        t0 = time.time()

        for ws_name in wb.sheetnames:
            if ws_name in self.exceptions:
                continue
            ws = wb[ws_name]

            # Backup sheet
            backup = wb.copy_worksheet(ws)
            backup.title = f"{ws_name}_backup"

            # Load into DataFrame
            df = pd.read_excel(
                file_path,
                sheet_name=ws_name,
                header=fr - 2,
                dtype=str
            )

            # Validate mock values if enabled
            if self.params['enable_mock_check'] == 1:
                mc = self.params['mock_col'] - 1
                invalid = ~df.iloc[:, mc].isin(['2', '3'])
                if invalid.any():
                    rows = df.index[invalid].tolist()
                    raise ValueError(f"Incorrect mock values in rows {rows} of sheet '{ws_name}'")

            # DeltaDelete: old (2) not in new (3)
            df = self._delta_delete(df)

            # DeltaNew: new (3) not in old (2)
            df = self._delta_new(df)

            # DeltaChange: compare adjacent old-new pairs
            drop_idx, to_color = self._delta_change(df, ws)

            # Drop identical new rows
            if drop_idx:
                df.drop(index=drop_idx, inplace=True)
                df.reset_index(drop=True, inplace=True)

            # Remap highlight coordinates after drops
            all_drops = sorted(drop_idx)
            remapped = []
            for (r, c) in to_color:
                offset = sum(1 for dr in all_drops if dr < r)
                remapped.append((r - offset, c))
            to_color = remapped

            # Clear existing data rows
            num_old = ws.max_row - fr + 1
            if num_old > 0:
                ws.delete_rows(fr, amount=num_old)
            # Write updated rows
            for row in dataframe_to_rows(df, index=False, header=False):
                ws.append(row)

            # Restore formatting
            backup_ws = wb[f"{ws_name}_backup"]
            self._restore_sheet_format(ws, backup_ws, fr)

            # Highlight changes in pink
            for r, c in to_color:
                ws.cell(row=fr + r, column=c + 1).fill = self._pink

        wb.save(file_path)
        print(f" → file done in {time.time()-t0:.2f}s")

    def _delta_delete(self, df: pd.DataFrame) -> pd.DataFrame:
        p = self.params
        mc = p['mock_col'] - 1
        kc = p['key_col'] - 1
        sc = p['status_prev_col'] - 1
        dc = p['delta_col'] - 1
        rng = df if p['record_limit'] == 0 else df.iloc[:p['record_limit']]
        # Collect new (mock=3) keys
        new_keys = set(rng.loc[rng.iloc[:, mc] == '3', df.columns[kc]])
        # Mark old=2 not in new, skip 'Add' rows
        mask = (
            (df.iloc[:, mc] == '2') &
            (~df.iloc[:, kc].isin(new_keys)) &
            (df.iloc[:, sc] != 'Add')
        )
        df.loc[mask, df.columns[dc]] = 'Delete'
        return df

    def _delta_new(self, df: pd.DataFrame) -> pd.DataFrame:
        p = self.params
        mc = p['mock_col'] - 1
        kc = p['key_col'] - 1
        dc = p['delta_col'] - 1
        rng = df if p['record_limit'] == 0 else df.iloc[:p['record_limit']]
        # Collect old (mock=2) keys
        old_keys = set(rng.loc[rng.iloc[:, mc] == '2', df.columns[kc]])
        # Mark new=3 not in old
        mask = (df.iloc[:, mc] == '3') & (~df.iloc[:, kc].isin(old_keys))
        df.loc[mask, df.columns[dc]] = 'New'
        return df

    def _delta_change(self, df: pd.DataFrame, ws) -> tuple[list[int], list[tuple[int,int]]]:
        p = self.params
        mc = p['mock_col'] - 1
        kc = p['key_col'] - 1
        dc = p['delta_col'] - 1
        hdr = p['header_row']
        n = len(df)

        # Identify adjacent old-new pairs (new=3, previous old=2)
        is_new = df.iloc[:, mc] == '3'
        is_old = df.iloc[:, mc].shift(1) == '2'
        same_key = df.iloc[:, kc] == df.iloc[:, kc].shift(1)
        pair = is_new & is_old & same_key

        # Determine which columns to compare
        start = p['start_compare_col'] - 1
        valid = []
        for j in range(start, df.shape[1]):
            h = str(ws.cell(row=hdr, column=j + 1).value or '').lower()
            if any(x in h for x in ('__', 'tobe', 'to be', 'to-be', 'to_be')):
                continue
            valid.append(j)

        drop_idx, to_color = [], []
        # Loop bottom-up
        for i in range(n - 1, 0, -1):
            if not pair.iat[i]:
                continue
            new_vals = df.iloc[i, valid].fillna('').astype(str)
            old_vals = df.iloc[i - 1, valid].fillna('').astype(str)
            diffs = new_vals != old_vals
            if not diffs.any():
                drop_idx.append(i)
            else:
                df.iat[i, dc] = 'Change'
                df.iat[i - 1, dc] = 'Change'
                # Collect cells to highlight
                for idx, changed in enumerate(diffs):
                    if changed:
                        col = valid[idx]
                        to_color.extend([(i, col), (i - 1, col)])
        return drop_idx, to_color

    def _restore_sheet_format(self, live_ws, backup_ws, first_row: int):
        # Copy column widths and row heights
        for col_letter, dim in backup_ws.column_dimensions.items():
            if dim.width:
                live_ws.column_dimensions[col_letter].width = dim.width
        for row_idx, dim in backup_ws.row_dimensions.items():
            if dim.height:
                live_ws.row_dimensions[row_idx].height = dim.height

        # Copy header styles
        hdr = first_row - 1
        for cell in backup_ws[hdr]:
            tgt = live_ws.cell(row=hdr, column=cell.column)
            for attr in ('font', 'border', 'fill', 'number_format', 'protection', 'alignment'):
                try:
                    setattr(tgt, attr, getattr(cell, attr))
                except:
                    pass

        # Apply data font and borders
        data_font = Font(name="Leelawadee", size=10)
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in live_ws.iter_rows(min_row=first_row, max_row=live_ws.max_row,
                                     min_col=1, max_col=live_ws.max_column):
            for cell in row:
                cell.font = data_font
                cell.border = border

if __name__ == '__main__':
    dp = DeltaProcessor(master_path="Master.xlsx")
    dp.run()
