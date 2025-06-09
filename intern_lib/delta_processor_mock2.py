# delta_processor_mock2.py

import os
import time
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side

class DeltaProcessor:
    """
    DeltaProcessor to apply delta logic (Delete, New, Change) exactly as in the VBA:
      1) DeltaDelete
      2) DeltaNew
      3) DeltaChange

    Backups each sheet, then rebuilds it and applies pink highlights.
    """

    def __init__(self, master_path: str):
        if not os.path.isfile(master_path):
            raise FileNotFoundError(f"Master workbook not found: {master_path}")
        self.master_path   = master_path
        self.params        = {}
        self.file_list     = None
        self.exceptions    = []
        self.sort_required = False
        self._pink         = PatternFill(start_color="F2CEF0",
                                         end_color="F2CEF0",
                                         fill_type="solid")
        self._load_master()

    def _load_master(self):
        p = pd.read_excel(self.master_path,
                          sheet_name="Parameters",
                          header=None, index_col=0)[1]
        p.index = p.index.astype(str).str.strip()
        self.params = {
            'first_row':         int(p['firstRowOfData']),
            'mock_col':          int(p['mockColumn']),
            'delta_col':         int(p['deltaIndicatorColumn']),
            'key_col':           int(p['concatenatedKeyColumn']),
            'status_prev_col':   int(p['statusFromPreviousMockColumn']),
            'record_limit':      int(p['recordLimit']),
            'header_row':        int(p['headerRow']),
            'start_compare_col': int(p['startColumnCheckData'])
        }
        self.file_list = pd.read_excel(self.master_path, sheet_name="File List")
        exc = pd.read_excel(self.master_path,
                            sheet_name="Exception Sheet Name").iloc[:,0]
        self.exceptions = exc.dropna().astype(str).str.strip().tolist()
        sr = pd.read_excel(self.master_path,
                           sheet_name="Sort Record",
                           header=None).iloc[1,0]
        self.sort_required = str(sr).strip().lower() == 'yes'

    def run(self):
        global_start = time.time()
        for _, row in self.file_list.iterrows():
            path, fname = row.iloc[0], row.iloc[1]
            full_path   = os.path.join(path, fname)
            if isinstance(path, str) and isinstance(fname, str) and os.path.exists(full_path):
                self._process_file(full_path)
        total = time.time() - global_start
        m, s = divmod(int(total), 60)
        print(f"\nALL WORKBOOKS DONE in {m}m {s}s")

    def _process_file(self, file_path: str):
        print(f"\nProcessing {os.path.basename(file_path)}")
        wb    = load_workbook(file_path)
        fr    = self.params['first_row']
        t0    = time.time()
        
        for ws_name in wb.sheetnames:
            if ws_name in self.exceptions:
                continue

            print(f" Sheet “{ws_name}”:")
            print(f" Start Time : {datetime.datetime.now()}")
            
            sheet_start = time.time()
            ws          = wb[ws_name]

            # 1) Backup
            t1 = time.time()
            backup = wb.copy_worksheet(ws)
            backup.title = f"{ws_name}_backup"
            print(f"  • backup: {time.time()-t1:.2f}s")

            # 2) Load into DataFrame
            t2 = time.time()
            df = pd.read_excel(
                file_path,
                sheet_name=ws_name,
                header=fr - 2,
                dtype=str
            )
            print(f"  • load df: {time.time()-t2:.2f}s")

            # 2.5) sort by [key, mock]
            t25 = time.time()
            kc = self.params['key_col'] - 1
            mc = self.params['mock_col'] - 1
            df.sort_values(by=[df.columns[kc], df.columns[mc]], inplace=True)
            df.reset_index(drop=True, inplace=True)
            print(f"  • sort: {time.time()-t25:.2f}s")

            # 4) DeltaDelete
            t4 = time.time()
            df = self._delta_delete(df)
            print(f"  • delta_delete: {time.time()-t4:.2f}s")

            # 5) DeltaNew
            t5 = time.time()
            df = self._delta_new(df)
            print(f"  • delta_new: {time.time()-t5:.2f}s")

            # 6) DeltaChange
            t6 = time.time()
            drop_idx, to_color = self._delta_change(df, ws)
            print(f"  • delta_change: {time.time()-t6:.2f}s")

            # 7) Drop identical mock-2 & reset
            t7 = time.time()
            if drop_idx:
                df.drop(index=drop_idx, inplace=True)
                df.reset_index(drop=True, inplace=True)
            print(f"  • drop rows: {time.time()-t7:.2f}s")

            # 7b) Remap the to_color coordinates
            #    so that row‐indices line up after the deletions
            all_drops = sorted(drop_idx)
            remapped = []
            for (r, c) in to_color:
                # count how many dropped rows were strictly above this row
                offset = sum(1 for dr in all_drops if dr < r)
                remapped.append((r - offset, c))
            to_color = remapped

            # 8) Clear & write in bulk -- optimized
            t8 = time.time()
            
            # delete everything from first_row to end
            num_old = ws.max_row - fr + 1
            if num_old > 0:
                ws.delete_rows(fr, amount=num_old)

            # now write new rows in one pass
            for row in dataframe_to_rows(df, index=False, header=False):
                ws.append(row)

            print(f"  • write sheet: {time.time()-t8:.2f}s")  
            
            # 8b) restore formatting
            t8b = time.time()
            backup_ws = wb[f"{ws_name}_backup"]
            self._restore_sheet_format(ws, backup_ws, fr)
            print(f"  • restore format: {time.time()-t8b:.2f}s")
                      
            # 9) Highlight “Change”
            t9 = time.time()
            for r, c in to_color:
                ws.cell(row=fr+r, column=c+1).fill = self._pink
            print(f"  • highlight: {time.time()-t9:.2f}s")

        wb.save(file_path)
        
        print(f" → file done in {time.time()-t0:.2f}s")

    def _delta_delete(self, df: pd.DataFrame) -> pd.DataFrame:
        p  = self.params
        mc = p['mock_col'] - 1
        kc = p['key_col']  - 1
        sc = p['status_prev_col'] - 1
        dc = p['delta_col'] - 1
        rng      = df if p['record_limit']==0 else df.iloc[:p['record_limit']]
        new_keys = set(rng.loc[rng.iloc[:,mc]=='2', df.columns[kc]])
        mask     = ((df.iloc[:,mc]=='1') &
                    (~df.iloc[:,kc].isin(new_keys)) &
                    (df.iloc[:,sc] != 'Add'))
        df.loc[mask, df.columns[dc]] = 'Delete'
        return df

    def _delta_new(self, df: pd.DataFrame) -> pd.DataFrame:
        p  = self.params
        mc = p['mock_col'] - 1
        kc = p['key_col']  - 1
        dc = p['delta_col'] - 1
        rng      = df if p['record_limit']==0 else df.iloc[:p['record_limit']]
        old_keys = set(rng.loc[rng.iloc[:,mc]=='1', df.columns[kc]])
        mask     = (df.iloc[:,mc]=='2') & (~df.iloc[:,kc].isin(old_keys))
        df.loc[mask, df.columns[dc]] = 'New'
        return df

    def _delta_change(self, df: pd.DataFrame, ws) -> tuple[list[int], list[tuple[int,int]]]:
        p   = self.params
        mc  = p['mock_col'] - 1
        kc  = p['key_col']  - 1
        dc  = p['delta_col'] - 1
        hdr = p['header_row']
        n   = len(df)

        is_new = df.iloc[:,mc]=='2'
        is_old = df.iloc[:,mc].shift(1)=='1'
        same   = df.iloc[:,kc]==df.iloc[:,kc].shift(1)
        pair   = is_new & is_old & same

        start = p['start_compare_col'] - 1
        valid = []
        for j in range(start, df.shape[1]):
            h = str(ws.cell(row=hdr, column=j+1).value or '').lower()
            if any(x in h for x in ('__','tobe','to be','to-be','to_be')):
                continue
            valid.append(j)

        drop_idx, to_color = [], []
        for i in range(n-1, 0, -1):
            if not pair.iat[i]:
                continue
            new_vals = df.iloc[i,   valid].fillna('').astype(str)
            old_vals = df.iloc[i-1, valid].fillna('').astype(str)
            diffs    = new_vals != old_vals
            if not diffs.any():
                drop_idx.append(i)
            else:
                df.iat[i,   dc] = 'Change'
                df.iat[i-1, dc] = 'Change'
                for idx, ch in enumerate(diffs):
                    if ch:
                        col = valid[idx]
                        to_color.extend([(i,   col), (i-1, col)])
        return drop_idx, to_color

    # def _restore_sheet_format(self, live_ws, backup_ws, first_row):
    #     # 1) Column widths
    #     for col_letter, dim in backup_ws.column_dimensions.items():
    #         if dim.width is not None:
    #             live_ws.column_dimensions[col_letter].width = dim.width

    #     # 2) Row heights
    #     for row_idx, dim in backup_ws.row_dimensions.items():
    #         if dim.height is not None:
    #             live_ws.row_dimensions[row_idx].height = dim.height

    #     # 3) Header‐row cell styles (guarded)
    #     hdr = first_row - 1
    #     for cell in backup_ws[hdr]:
    #         new_cell = live_ws.cell(row=hdr, column=cell.column)
    #         # try to copy each style attribute; skip if unhashable
    #         for attr in ('font', 'border', 'fill', 'number_format',
    #                     'protection', 'alignment'):
    #             try:
    #                 setattr(new_cell, attr, getattr(cell, attr))
    #             except TypeError:
    #                 # e.g. unhashable StyleProxy → just skip
    #                 pass
    
    def _restore_sheet_format(self, live_ws, backup_ws, first_row: int):
        """
        1) Copy column widths & row heights
        2) Copy header styles
        3) Apply Leelawadee font size 10 and thin border to loaded data
        """
        # 1) Column widths
        for col_letter, dim in backup_ws.column_dimensions.items():
            if dim.width:
                live_ws.column_dimensions[col_letter].width = dim.width

        # 2) Row heights
        for row_idx, dim in backup_ws.row_dimensions.items():
            if dim.height:
                live_ws.row_dimensions[row_idx].height = dim.height

        # 3) Header‐row styles
        hdr = first_row - 1
        for cell in backup_ws[hdr]:
            tgt = live_ws.cell(row=hdr, column=cell.column)
            for attr in (
                'font', 'border', 'fill',
                'number_format', 'protection', 'alignment'
            ):
                try:
                    setattr(tgt, attr, getattr(cell, attr))
                except:
                    pass

        # 4) Leelawadee font & thin border on data cells
        data_font = Font(name="Leelawadee", size=10)
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        max_row = live_ws.max_row
        max_col = live_ws.max_column
        for row in live_ws.iter_rows(
            min_row=first_row, max_row=max_row,
            min_col=1, max_col=max_col
        ):
            for cell in row:
                cell.font = data_font
                cell.border = border


if __name__ == '__main__':
    dp = DeltaProcessor(master_path="Master.xlsx")
    dp.run()
